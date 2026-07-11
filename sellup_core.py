"""
sellup_core.py
==============
Standalone core logic for the SellUp stock bulk-update automation.

This module is intentionally UI-free (no Streamlit imports) so it can be unit
tested and reused. The Streamlit front-end (app.py) is a thin wrapper on top.

It is a *separate* codebase from the Shopee automation. It shares only the
high-level idea (build a trusted match registry, then push quantities across
it); none of the Shopee matching rules apply here because SellUp uses clean
structured columns instead of Shopee's free-text Product/Variation names.

--------------------------------------------------------------------------
INPUT: SellUp "Dealer Inventory Bulk Update" export (INVENTORIES_*.xlsx)
--------------------------------------------------------------------------
One sheet per category: Smartphones, Tablets, Smartwatches, Game Consoles, Audio.
Row layout per sheet:
    row 1: title  "DEALER INVENTORY BULK UPDATE"
    row 2: instructions
    row 3: headers
    row 4+: data
Columns (15):
    A SKU ID | B Brand | C Model | D Specs | E Color |
    F New(Not Activated) Price | G New(Not Activated) Qty |
    H New(Activated) Price     | I New(Activated) Qty |
    J Excellent Price          | K Excellent Qty |
    L Good Price               | M Good Qty |
    N Fair Price               | O Fair Qty

--------------------------------------------------------------------------
INPUT: Masterlist (stock_report.xlsx) - single "Worksheet"
--------------------------------------------------------------------------
    Stock Type ID | Category(New/Used) | Brand | Model | Color | Total | ...stores
The Model field embeds storage + connectivity + region + condition markers.

--------------------------------------------------------------------------
MATCHING MODEL (differs from Shopee)
--------------------------------------------------------------------------
Both sides are normalised to a common key:  BRAND | MODEL_BASE | STORAGE | COLOUR
Condition routing (which SellUp Qty column a Masterlist row feeds):
    Category "New"  + model marker "NA"  -> New (Not Activated)
    Category "New"  + model marker "A"   -> New (Activated)
    Category "New"  + no marker          -> New (Not Activated)   [default, flagged]
    Category "Used"                      -> Excellent
Region-suffixed Masterlist SKUs (HK/US/AUS/...) merge into the base model
(their quantities are summed into the plain model).
A SellUp condition Qty is only written when that condition's PRICE cell is
populated and > 0 (SellUp treats an empty price as "not sold in this grade").
Only Qty is ever written; Price cells are never touched.
"""

from __future__ import annotations
import re
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from rapidfuzz import fuzz

# ----------------------------------------------------------------------------
# Configuration / vocabularies
# ----------------------------------------------------------------------------

SELLUP_HEADER_ROW = 3          # 1-indexed row that holds column headers
SELLUP_DATA_START = 4          # 1-indexed first data row

# SellUp condition columns -> (price_col_idx, qty_col_idx)  (0-indexed)
CONDITIONS = {
    "New (Not Activated)": (5, 6),
    "New (Activated)":     (7, 8),
    "Excellent":           (9, 10),
    "Good":                (11, 12),
    "Fair":                (13, 14),
}

# Which SellUp condition each Masterlist (category, marker) routes to.
# marker is "NA", "A" or "" (none)
def route_condition(category: str, marker: str) -> str:
    category = category.strip().upper()
    if category == "USED":
        return "Excellent"
    if category == "NEW":
        if marker == "A":
            return "New (Activated)"
        return "New (Not Activated)"   # NA or unmarked default
    return ""  # unknown category -> unmatched

# Region / market suffix tokens that should be stripped and merged into base.
REGION_TOKENS = {
    "HK", "US", "USA", "CN", "JP", "UAE", "IND", "INDIA", "INDO", "ID",
    "UK", "KR", "TW", "EU", "SG", "MY", "AUS", "AU", "TH", "PH", "VN",
    "DK", "DE", "FR", "IT", "ES", "CA", "GLOBAL", "ROW", "ME",
}

# Condition markers (New only): activation status.
MARKER_NA = "NA"
MARKER_A = "A"

# Connectivity tags. These are stripped from the model BASE (so bases align) and
# are instead captured as a separate `connectivity` key dimension for tablets and
# smartwatches (where Wi-Fi vs Cellular is a real product difference). For phones
# there is no Wi-Fi-only variant, so connectivity is intentionally ignored there.
CONNECTIVITY_TOKENS = {"5G", "4G", "3G", "LTE", "VOLTE",
                       "WIFI", "CELL", "CELLULAR", "GPS", "BLUETOOTH"}

# Regex marking a CELLULAR-capable variant (vs Wi-Fi/GPS-only = "NONCELL").
_CELL_RE = re.compile(r"CELLULAR|\bCELL\b|\bLTE\b|\b[345]G\b")
_MM_RE = re.compile(r"\b(\d{2})\s*MM\b")


def device_type(raw_brand: str, model: str, sheet: str = "") -> str:
    """Classify a row as 'tablet', 'watch', 'phone', 'audio' or 'other'.

    For SellUp rows the `sheet` name is authoritative; for Masterlist rows we
    infer from the (un-aliased) brand and model text.
    """
    s = _up(sheet)
    if s == "TABLETS":
        return "tablet"
    if s == "SMARTWATCHES":
        return "watch"
    if s == "SMARTPHONES":
        return "phone"
    if s == "AUDIO":
        return "audio"
    if s in ("GAME CONSOLES",):
        return "other"
    rb = _up(raw_brand)
    m = _up(model)
    if "WATCH" in rb or "WATCH" in m:
        return "watch"
    if re.search(r"\b(AIRPODS|BUDS|EARBUDS)\b", m) or re.match(r"^W[FH]-?\d", m):
        return "audio"
    if rb == "IPAD" or "TABLET" in rb:
        return "tablet"
    # tablet if PAD / TAB / MATEPAD appears as a word anywhere in the model
    if re.search(r"\b(TAB|PAD|MATEPAD)\b", m):
        return "tablet"
    return "other"


def connectivity_class(text: str, dtype: str) -> str:
    """WIFI/GPS-only -> 'NONCELL'; anything cellular -> 'CELL'.
    Only meaningful for tablets & watches; returns '' otherwise."""
    if dtype not in ("tablet", "watch"):
        return ""
    return "CELL" if _CELL_RE.search(_up(text)) else "NONCELL"


def case_size(text: str, dtype: str) -> str:
    """Watch case size like '44MM' (a per-watch discriminator). '' otherwise."""
    if dtype != "watch":
        return ""
    m = _MM_RE.search(_up(text))
    return f"{m.group(1)}MM" if m else ""


# ---------------------------------------------------------------------------
# AUDIO parser  (AirPods / Buds / Sony) — names differ by connector/generation
# ---------------------------------------------------------------------------
# Rules derived from the data:
#   - "Type C" == "USB-C"        (connector synonyms)
#   - "(MagSafe)" is noise
#   - AirPods Pro/Max with no generation number => generation 1
#   - "with Active Noise Cancellation" is a real product distinction (kept)
#   - connector is a real distinction when BOTH sides state it (Max Lightning
#     vs Max USB-C), but optional when one side omits it (Pro 1).

def parse_audio(model: str) -> tuple:
    """Return (core_key, connector). core_key is an order-independent string."""
    m = _up(model)
    m = re.sub(r"\(.*?\)", " ", m)                       # drop (MagSafe)
    for syn in ("TYPE C", "TYPE-C", "USB-C", "USB C"):
        m = m.replace(syn, "USBC")
    anc = "ANC" if "ACTIVE NOISE CANCELLATION" in m else ""
    m = m.replace("WITH ACTIVE NOISE CANCELLATION", " ")
    m = re.sub(r"\bMAGSAFE\b", " ", m)
    m = re.sub(r"^GALAXY\s+", " ", m)                    # Samsung brand word
    m = re.sub(r"(BUDS)(\d)", r"\1 \2", m)               # Buds2 -> Buds 2
    connector = ""
    if "LIGHTNING" in m:
        connector = "LIGHTNING"; m = m.replace("LIGHTNING", " ")
    elif "USBC" in m:
        connector = "USBC"; m = m.replace("USBC", " ")
    toks = [t for t in m.split() if t and t not in FLAVOUR_TOKENS]
    if "AIRPODS" in toks and not any(any(c.isdigit() for c in t) for t in toks):
        toks.append("1")                                 # implicit generation 1
    if anc:
        toks.append("ANC")
    return " ".join(sorted(toks)), connector


# ---------------------------------------------------------------------------
# WATCH parser — SellUp "Watch SE 2 Aluminium" + "44mm, GPS" + "Midnight"
#   vs Masterlist "SE 2 44MM GPS MIDNIGHT ALUMINIUM CASE ... SPORT BAND ..."
# ---------------------------------------------------------------------------
_WATCH_MATERIALS = ("STAINLESS STEEL", "ALUMINIUM", "ALUMINUM", "TITANIUM")
_WATCH_LINE_RE = re.compile(r"\b(SE|SERIES|ULTRA)\b")


def _material(text: str) -> tuple:
    t = _up(text).replace("ALUMINUM", "ALUMINIUM")
    if "STAINLESS STEEL" in t or "STEEL" in t:
        return "STEEL", "STEEL"
    if "TITANIUM" in t:
        return "TITANIUM", "TITANIUM"
    if "ALUMINIUM" in t:
        return "ALUMINIUM", "ALUMINIUM"
    return "", ""


def _watch_line_gen(m: str) -> tuple:
    lg = re.match(r"(SE|SERIES|ULTRA)\s+(\d+)(?!\d*\s*MM)", m)  # gen, not the size
    if lg:
        return lg.group(1), lg.group(2)
    l = _WATCH_LINE_RE.search(m)
    if l:
        return l.group(1), "1"        # e.g. "Ultra" with no number = gen 1
    return None, None


def parse_watch_ml(model: str):
    """Parse a Masterlist watch model into a canonical signature dict."""
    m = _up(model)
    line, gen = _watch_line_gen(m)
    if not line:
        return None
    size = _MM_RE.search(m)
    size = f"{size.group(1)}MM" if size else ""
    conn = "CELL" if re.search(r"\bCELL\b|\bLTE\b", m) else ("GPS" if "GPS" in m else "")
    mat, matword = _material(m)
    case_color = ""
    if matword:
        before = m.split(matword)[0].split()
        drop = {line, gen, size, "GPS", "CELL", "LTE", "CELLULAR", "+", "WITH"}
        cc = [t for t in before if t not in drop and not t.endswith("MM")]
        if cc and cc[0] == line:
            cc = cc[1:]
        if cc and cc[0] == gen:
            cc = cc[1:]
        case_color = norm_colour(" ".join(cc))
    return {"line": line, "gen": gen, "size": size, "conn": conn,
            "mat": mat, "colour": case_color}


def parse_watch_sellup(model: str, specs: str, colour: str):
    """Parse a SellUp watch row (Model + Specs + Colour) into the same sig."""
    m = re.sub(r"^WATCH\s+", "", _up(model))
    line, gen = _watch_line_gen(m)
    if not line:
        return None
    mat, _ = _material(m)
    size = _MM_RE.search(_up(specs))
    size = f"{size.group(1)}MM" if size else ""
    s = _up(specs)
    conn = "CELL" if re.search(r"CELL|LTE", s) else ("GPS" if "GPS" in s else "")
    # SellUp colour often carries the material word, e.g. "Natural Titanium"
    cc = _up(colour)
    for w in ("STAINLESS STEEL", "TITANIUM", "ALUMINIUM", "ALUMINUM", "STEEL"):
        cc = cc.replace(w, " ")
    case_color = norm_colour(cc)
    return {"line": line, "gen": gen, "size": size, "conn": conn,
            "mat": mat, "colour": case_color}


def watch_model_base(sig: dict) -> str:
    return f"{sig['line']} {sig['gen']} {sig['mat']}".strip()

# Bundle / flavour tags used on New phones in the Masterlist (SellUp has none).
FLAVOUR_TOKENS = {
    "PRIMARY", "TELCO", "FREEBIES", "FREEB", "WFREEB", "WITHFREEBIES",
    "BASIC", "SET", "FULLSET", "LOCALSET", "EXPORT", "EXPORTSET",
}

# A Samsung/vendor model-code token, e.g. "S948B", "A576B", "F966", "S908E",
# and the hyphenated "5G-S948B" form. Stripped from the base.
MODELCODE_RE = re.compile(r"^(?:\dG-)?[A-Z]?\d{3,}[A-Z0-9]*$")

# Brand synonym normalisation (SellUp brand -> canonical) and Masterlist quirks.
BRAND_ALIASES = {
    "ONE PLUS": "ONEPLUS",
    "VIVO": "VIVO",
    "APPLE WATCH": "APPLE",
    "SAMSUNG WATCH": "SAMSUNG",
    "GOOGLE WATCH": "GOOGLE",
    "OPPO WATCH": "OPPO",
    "HONOR TABLET": "HONOR",
    # Masterlist "brand" is often a product-line label; normalise the Apple lines:
    "IPHONE": "APPLE",
    "IPAD": "APPLE",
}

# Colour normalisation aliases (applied after upper+trim).
COLOUR_ALIASES = {
    "SPACE GREY": "SPACE GRAY",
    "GREY": "GRAY",
    "MIDNIGHT BLACK": "MIDNIGHT",
}

STORAGE_RE = re.compile(r"\b(\d+(?:\.\d+)?)\s*(TB|GB)\b")
RAMSTOR_RE = re.compile(r"\b(\d+)\s*/\s*(\d+(?:\.\d+)?)\s*(TB|GB)\b")   # 12/128GB
STORRAM_RE = re.compile(r"\b(\d+(?:\.\d+)?)\s*(TB|GB)\s*/\s*(\d+)\b")  # 128GB/12

# Thresholds for the fuzzy fallback on the model base.
FUZZY_HIGH = 90    # >= -> auto (High confidence)
FUZZY_REVIEW = 78  # >= -> Match Review (needs a human)

# Tier / line keywords that distinguish products sharing a model number
# (e.g. iPhone 15 vs 15 PRO vs 15 PRO MAX vs 15 PLUS).
TIER_TOKENS = {
    "PRO", "MAX", "PLUS", "MINI", "FE", "ULTRA", "AIR", "LITE", "NEO",
    "NOTE", "EDGE", "PRIME", "ACTIVE", "FOLD", "FLIP", "SE", "GT", "POCO",
    "REDMI", "TURBO", "POWER", "STAR", "MAGIC", "NORD", "REALME", "XL",
}


def model_signature(model_base: str) -> tuple:
    """
    A product signature that must match EXACTLY for two model bases to be the
    same product. Guards against fuzzy string similarity conflating different
    generations, e.g. '11 PRO MAX' vs '12 PRO MAX' (one digit apart, ~90%
    similar, but a completely different phone).

    Returns (frozenset of digit-bearing tokens, frozenset of tier tokens).
    """
    toks = model_base.split()
    nums = frozenset(t for t in toks if any(ch.isdigit() for ch in t))
    tiers = frozenset(t for t in toks if t in TIER_TOKENS)
    return (nums, tiers)


# ----------------------------------------------------------------------------
# Normalisation helpers
# ----------------------------------------------------------------------------

def _up(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().upper())


def norm_brand(b: str) -> str:
    b = _up(b)
    return BRAND_ALIASES.get(b, b)


def norm_colour(c: str) -> str:
    c = _up(c)
    return COLOUR_ALIASES.get(c, c)


def extract_storage(text: str) -> str:
    """Return a single canonical storage token e.g. '128GB' or '' if none.

    Prefers the STORAGE part of a RAM/STORAGE combo (e.g. 12/256GB -> 256GB).
    """
    t = _up(text)
    combo = RAMSTOR_RE.findall(t)          # 12/128GB -> storage = 128GB
    if combo:
        ram, stor, unit = combo[0]
        return f"{stor}{unit}"
    combo2 = STORRAM_RE.findall(t)          # 128GB/12 -> storage = 128GB
    if combo2:
        stor, unit, ram = combo2[0]
        return f"{stor}{unit}"
    m = STORAGE_RE.findall(t)
    if m:
        num, unit = m[0]
        return f"{num}{unit}"
    return ""


def detect_marker(model_tokens: list[str]) -> str:
    """Detect activation marker NA / A among model tokens (New category only)."""
    toks = set(model_tokens)
    if MARKER_NA in toks:
        return "NA"
    if MARKER_A in toks:
        return "A"
    return ""


# Tokens that are never part of the model identity for KEY purposes.
# (connectivity/flavour tokens are kept out of the base so keys align with
#  SellUp's clean Model column, which never contains them.)
NOISE_TOKENS = (REGION_TOKENS | CONNECTIVITY_TOKENS | FLAVOUR_TOKENS
                | {MARKER_NA, MARKER_A})

# Leftover RAM fragments like "/12", "12/", or a stray "12" left after storage
# removal in a "RAM/STORAGE" combo (e.g. "12/128GB" -> storage 128GB, drop 12).
_RAM_FRAG_RE = re.compile(r"(?<!\S)\d{1,3}\s*/|/\s*\d{1,3}(?!\S)|(?<!\S)/(?!\S)")


def masterlist_model_base(model: str) -> tuple[str, str, str]:
    """
    Parse a Masterlist Model string.
    Returns (model_base, storage, marker).
    model_base = model with storage, region, connectivity, flavour, vendor
    model-codes and activation markers removed.
    """
    raw = _up(model)
    storage = extract_storage(raw)
    marker = detect_marker(raw.split())     # detect NA / A before stripping
    # remove storage substrings (both RAM/storage orders first, then plain)
    base = RAMSTOR_RE.sub(" ", raw)     # 12/128GB
    base = STORRAM_RE.sub(" ", base)    # 128GB/12
    base = STORAGE_RE.sub(" ", base)
    # clean leftover RAM fragments / stray slashes
    base = _RAM_FRAG_RE.sub(" ", base)
    base = base.replace("/", " ")
    toks = base.split()
    toks = [t for t in toks
            if t not in NOISE_TOKENS and not MODELCODE_RE.match(t)]
    base = re.sub(r"\s+", " ", " ".join(toks)).strip()
    return base, storage, marker


def sellup_model_base(brand: str, model: str) -> str:
    """Normalise a SellUp Model into a base comparable to the Masterlist base."""
    m = _up(model)
    # Drop a leading brand word if repeated in the model (e.g. "Apple iPhone 4")
    m = re.sub(
        r"^(APPLE|SAMSUNG|GOOGLE|XIAOMI|HUAWEI|HONOR|OPPO|VIVO|ONEPLUS|"
        r"NOTHING|ASUS|INFINIX|SONY|MICROSOFT|NINTENDO|AMAZON|DJI|OCULUS|"
        r"LENOVO|MSI|ROG|VALVE)\s+",
        "", m,
    )
    # Apple phones: Masterlist drops the word IPHONE (brand carries it)
    m = re.sub(r"^IPHONE\s+", "", m)
    m = re.sub(r"^IPAD\s+", "IPAD ", m)  # keep iPad token, matched below
    # drop connectivity/region noise for symmetry with the Masterlist base
    toks = [t for t in m.split()
            if t not in CONNECTIVITY_TOKENS and t not in REGION_TOKENS]
    m = re.sub(r"\s+", " ", " ".join(toks)).strip()
    return m


# ----------------------------------------------------------------------------
# Data records
# ----------------------------------------------------------------------------

@dataclass
class MasterRow:
    stock_id: str
    category: str          # New / Used
    brand: str             # normalised
    model_base: str        # normalised
    storage: str
    colour: str
    marker: str            # NA / A / ""
    condition: str         # routed SellUp condition
    qty: int
    raw_model: str = ""
    raw_colour: str = ""   # original colour text from the Masterlist
    connectivity: str = "" # CELL / NONCELL / GPS / "" (tablets & watches)
    size_mm: str = ""      # watch case size, e.g. "44MM"
    dtype: str = "other"   # phone / tablet / watch / audio / other
    connector: str = ""    # audio only: LIGHTNING / USBC / ""

    @property
    def key(self) -> tuple:
        return (self.brand, self.model_base, self.storage, self.colour,
                self.connectivity, self.size_mm)


@dataclass
class SellUpRow:
    sheet: str
    excel_row: int         # 1-indexed row in the sheet
    sku_id: str
    brand: str             # normalised
    model_base: str        # normalised
    storage: str
    colour: str
    prices: dict = field(default_factory=dict)   # condition -> price value (raw)
    qtys: dict = field(default_factory=dict)      # condition -> current qty (raw)
    raw_colour: str = ""   # original colour text from the SellUp file
    raw_model: str = ""    # original Model text (for display; base is internal)
    connectivity: str = "" # CELL / NONCELL / GPS / "" (tablets & watches)
    size_mm: str = ""      # watch case size, e.g. "44MM"
    dtype: str = "other"   # phone / tablet / watch / audio / other
    connector: str = ""    # audio only: LIGHTNING / USBC / ""

    @property
    def key(self) -> tuple:
        return (self.brand, self.model_base, self.storage, self.colour,
                self.connectivity, self.size_mm)

    def price_active(self, condition: str) -> bool:
        v = self.prices.get(condition)
        try:
            return v not in (None, "") and float(v) > 0
        except (TypeError, ValueError):
            return False


# ----------------------------------------------------------------------------
# Loaders
# ----------------------------------------------------------------------------

def load_masterlist(path_or_stream) -> list[MasterRow]:
    wb = openpyxl.load_workbook(path_or_stream, read_only=True, data_only=True)
    ws = wb["Worksheet"] if "Worksheet" in wb.sheetnames else wb.worksheets[0]
    out: list[MasterRow] = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or r[0] in (None, ""):
            continue
        category = _up(r[1])
        if category not in ("NEW", "USED"):
            continue
        brand = norm_brand(r[2])
        base, storage, marker = masterlist_model_base(r[3])
        colour = norm_colour(r[4])
        try:
            qty = int(r[5]) if r[5] not in (None, "") else 0
        except (TypeError, ValueError):
            qty = 0
        cond = route_condition(category, marker)
        dtype = device_type(r[2], r[3])
        raw_model = _up(r[3])
        conn = connectivity_class(raw_model, dtype)
        size = case_size(raw_model, dtype)
        connector = ""
        # --- category-specific canonicalisation ---
        if dtype == "audio":
            base, connector = parse_audio(raw_model)
            storage = ""; conn = ""; size = ""
        elif dtype == "watch":
            wsig = parse_watch_ml(raw_model)
            if wsig:
                base = watch_model_base(wsig)
                colour = wsig["colour"] or colour
                storage = ""; conn = wsig["conn"]; size = wsig["size"]
        out.append(MasterRow(
            stock_id=str(r[0]), category=category.title(), brand=brand,
            model_base=base, storage=storage, colour=colour, marker=marker,
            condition=cond, qty=qty, raw_model=raw_model,
            raw_colour=str(r[4] or "").strip(),
            connectivity=conn, size_mm=size, dtype=dtype, connector=connector,
        ))
    return out


def load_sellup(path_or_stream) -> tuple[dict[str, list[SellUpRow]], "openpyxl.Workbook"]:
    """
    Returns (rows_by_sheet, workbook).
    The workbook is loaded in read/write mode so app.py can write Qty back and
    re-save while preserving the exact SellUp template.
    """
    wb = openpyxl.load_workbook(path_or_stream)   # NOT read_only -> writable
    rows_by_sheet: dict[str, list[SellUpRow]] = {}
    for ws in wb.worksheets:
        sheet_rows: list[SellUpRow] = []
        for excel_row in range(SELLUP_DATA_START, ws.max_row + 1):
            sku = ws.cell(row=excel_row, column=1).value
            if sku in (None, ""):
                continue
            brand = norm_brand(ws.cell(row=excel_row, column=2).value)
            model = ws.cell(row=excel_row, column=3).value
            specs = ws.cell(row=excel_row, column=4).value
            colour_raw = ws.cell(row=excel_row, column=5).value
            colour = norm_colour(colour_raw)
            base = sellup_model_base(brand, model)
            storage = extract_storage(specs)
            dtype = device_type(brand, model, ws.title)
            conn = connectivity_class(specs, dtype)
            size = case_size(specs, dtype)
            connector = ""
            # --- category-specific canonicalisation (mirror the Masterlist) ---
            if dtype == "audio":
                base, connector = parse_audio(model)
                storage = ""; conn = ""; size = ""
            elif dtype == "watch":
                wsig = parse_watch_sellup(model, specs, colour_raw)
                if wsig:
                    base = watch_model_base(wsig)
                    colour = wsig["colour"] or colour
                    storage = ""; conn = wsig["conn"]; size = wsig["size"]
            prices, qtys = {}, {}
            for cond, (pcol, qcol) in CONDITIONS.items():
                prices[cond] = ws.cell(row=excel_row, column=pcol + 1).value
                qtys[cond] = ws.cell(row=excel_row, column=qcol + 1).value
            sheet_rows.append(SellUpRow(
                sheet=ws.title, excel_row=excel_row, sku_id=str(sku),
                brand=brand, model_base=base, storage=storage,
                colour=colour, prices=prices, qtys=qtys,
                raw_colour=str(colour_raw or "").strip(),
                raw_model=str(model or "").strip(),
                connectivity=conn, size_mm=size, dtype=dtype,
                connector=connector,
            ))
        rows_by_sheet[ws.title] = sheet_rows
    return rows_by_sheet, wb


# ----------------------------------------------------------------------------
# Matching
# ----------------------------------------------------------------------------

@dataclass
class MatchResult:
    sellup: SellUpRow
    condition: str                 # SellUp condition column being filled
    master_rows: list[MasterRow]   # contributing masterlist rows (merged)
    qty: int
    confidence: str                # "Exact" / "High" / "Review" / "None"
    score: float = 100.0
    note: str = ""


def build_master_index(master: list[MasterRow]):
    """Index masterlist rows by exact key and by (brand, storage, colour) for fuzzy."""
    by_key: dict[tuple, list[MasterRow]] = defaultdict(list)
    by_bsc: dict[tuple, list[MasterRow]] = defaultdict(list)
    for m in master:
        if not m.condition or excluded_from_sync(m):
            continue  # Telco / freebies never contribute stock
        by_key[m.key].append(m)
        by_bsc[(m.brand, m.storage, m.colour,
                m.connectivity, m.size_mm)].append(m)
    return by_key, by_bsc


def match(master: list[MasterRow], rows_by_sheet: dict[str, list[SellUpRow]],
          crosswalk: dict | None = None):
    """
    For every SellUp row and every ACTIVE (priced) condition, find the
    contributing Masterlist rows and compute the quantity.

    crosswalk: authoritative {masterlist_stock_id: sellup_sku_id} map (from a
    known-good mapping file). These links are applied FIRST and take priority
    over fuzzy matching; the cells they claim are not re-matched.

    Returns (results, unmatched_master) where results is a list[MatchResult].
    """
    crosswalk = crosswalk or {}
    used_master_ids: set[str] = set()
    claimed_cells: set = set()          # (sku_id, condition) owned by crosswalk
    results: list[MatchResult] = []

    # ---- Phase A: authoritative crosswalk links ----
    if crosswalk:
        ml_by_id = {m.stock_id: m for m in master}
        su_by_sku = {s.sku_id: s for rr in rows_by_sheet.values() for s in rr}
        agg = defaultdict(lambda: [None, 0, []])
        for sid, sku in crosswalk.items():
            m = ml_by_id.get(sid)
            s = su_by_sku.get(sku)
            if not m or not s or not m.condition or excluded_from_sync(m):
                continue  # telco/freebies handled downstream, not synced
            cell = agg[(s.sku_id, m.condition)]
            cell[0] = s
            cell[2].append(m)
            if s.price_active(m.condition):
                cell[1] += m.qty
            used_master_ids.add(sid)
        for (sku, cond), (s, qty, mss) in agg.items():
            results.append(MatchResult(s, cond, mss, qty, "Exact", 100.0,
                                       note="crosswalk"))
            claimed_cells.add((sku, cond))

    # ---- Phase B: fuzzy/exact engine for everything not claimed ----
    engine_master = [m for m in master if m.stock_id not in used_master_ids]
    by_key, by_bsc = build_master_index(engine_master)

    for sheet, rows in rows_by_sheet.items():
        for s in rows:
            for cond in CONDITIONS:
                if not s.price_active(cond):
                    continue  # SellUp: empty/zero price -> skip this grade
                if (s.sku_id, cond) in claimed_cells:
                    continue  # crosswalk owns this cell

                # candidate masterlist rows must route to THIS condition.
                # For audio, connector must be compatible (equal, or the
                # Masterlist omits it): distinguishes Max Lightning vs Max USB-C
                # while still matching Pro 1 (Masterlist has no connector).
                exact = [m for m in by_key.get(s.key, [])
                         if m.condition == cond
                         and (s.dtype != "audio"
                              or not m.connector
                              or m.connector == s.connector)]
                if exact:
                    qty = sum(m.qty for m in exact)
                    for m in exact:
                        used_master_ids.add(m.stock_id)
                    results.append(MatchResult(
                        s, cond, exact, qty, "Exact", 100.0))
                    continue

                # fuzzy fallback: same brand+storage+colour AND same product
                # signature (model-number + tier tokens). The signature guard
                # prevents cross-generation matches (11 PRO MAX vs 12 PRO MAX).
                s_sig = model_signature(s.model_base)
                cands = [m for m in by_bsc.get(
                            (s.brand, s.storage, s.colour,
                             s.connectivity, s.size_mm), [])
                         if m.condition == cond
                         and model_signature(m.model_base) == s_sig]
                best, best_score = None, 0.0
                grouped: dict[str, list[MasterRow]] = defaultdict(list)
                for m in cands:
                    grouped[m.model_base].append(m)
                for mb, grp in grouped.items():
                    sc = fuzz.token_sort_ratio(s.model_base, mb)
                    if sc > best_score:
                        best, best_score = grp, sc
                if best and best_score >= FUZZY_HIGH:
                    qty = sum(m.qty for m in best)
                    for m in best:
                        used_master_ids.add(m.stock_id)
                    results.append(MatchResult(
                        s, cond, best, qty, "High", best_score,
                        note=f"fuzzy model '{best[0].model_base}'"))
                elif best and best_score >= FUZZY_REVIEW:
                    results.append(MatchResult(
                        s, cond, best, sum(m.qty for m in best), "Review",
                        best_score, note=f"fuzzy model '{best[0].model_base}'"))
                else:
                    results.append(MatchResult(
                        s, cond, [], 0, "None", 0.0,
                        note="no masterlist match"))

    unmatched_master = [m for m in master
                        if m.condition and m.stock_id not in used_master_ids]
    return results, unmatched_master


# ----------------------------------------------------------------------------
# Write-back
# ----------------------------------------------------------------------------

def apply_quantities(wb, results: list[MatchResult], only_confidences=("Exact", "High")):
    """
    Write computed Qty into the SellUp workbook, IN PLACE, preserving template.
    Only writes rows whose confidence is in `only_confidences` (locked matches).
    Returns number of cells written.
    """
    written = 0
    for res in results:
        if res.confidence not in only_confidences:
            continue
        ws = wb[res.sellup.sheet]
        _pcol, qcol = CONDITIONS[res.condition]
        ws.cell(row=res.sellup.excel_row, column=qcol + 1).value = res.qty
        written += 1
    return written


# ----------------------------------------------------------------------------
# Stats helper
# ----------------------------------------------------------------------------

def summarise(results: list[MatchResult]) -> dict:
    stats = defaultdict(int)
    for r in results:
        stats[r.confidence] += 1
    return dict(stats)


# ----------------------------------------------------------------------------
# Registry (Match Review) workbook builder  --  Shopee-styled design
# ----------------------------------------------------------------------------

# Palette lifted from the Shopee registry so the two workbooks look identical.
CLR_NAVY = "FF1F3864"    # structural columns (# and last), + Summary header/title
CLR_PEACH = "FFF4B183"   # source columns (SellUp side)
CLR_GOLD = "FFFFD966"    # decision / target columns (Masterlist side / actions)
CLR_GREY = "FF808080"    # footnote text

REGISTRY_TABS = ["Summary", "Locked Matches", "New Masterlist SKUs",
                 "Match Review", "Skipped (No Price)", "Not Selling in SellUp",
                 "Not on SellUp Yet"]

# Reviewer-decision options offered in the New Masterlist SKUs dropdown.
NM_DECISIONS = ["Linked", "Not on SellUp Yet", "Not Selling in SellUp",
                "Skipped (No Price)"]


def is_freebie(m: "MasterRow") -> bool:
    """A giveaway / gift-with-purchase item (never sold on SellUp)."""
    return "FREEB" in (m.raw_model or "").upper()


def is_telco(m: "MasterRow") -> bool:
    """A carrier/telco-bundled set. MM only sells Primary sets on SellUp, so
    Telco stock must NOT be matched or summed into a listing."""
    return "TELCO" in (m.raw_model or "").upper()


def excluded_from_sync(m: "MasterRow") -> bool:
    """Masterlist rows that must never contribute stock to a SellUp listing."""
    return is_freebie(m) or is_telco(m)


def _norm_id(v) -> str:
    """Normalise a stock-type id read back from Excel (may be int/float/text)."""
    if v in (None, ""):
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return str(v).strip()


def read_crosswalk(path_or_stream) -> dict:
    """Read an authoritative {masterlist_stock_id: sellup_sku_id} mapping from a
    'SellUp Data'-style workbook with 'Masterlist Stock Type ID' and
    'SellUp SKU ID' columns."""
    import openpyxl
    wb = openpyxl.load_workbook(path_or_stream, data_only=True)
    ws = wb.worksheets[0]
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

    def idx(name):
        for i, h in enumerate(hdr):
            if h.lower() == name.lower():
                return i
        return None
    ci, cj = idx("Masterlist Stock Type ID"), idx("SellUp SKU ID")
    out = {}
    if ci is None or cj is None:
        return out
    for r in ws.iter_rows(min_row=2, values_only=True):
        mid = _norm_id(r[ci])
        sku = str(r[cj]).strip() if r[cj] not in (None, "") else ""
        if mid and sku:
            out.setdefault(mid, sku)   # first mapping wins on dup ids
    return out


def read_registry_links(path_or_stream) -> dict:
    """Harvest ALL confirmed masterlist->SellUp links from a reviewed registry:
    the 'Locked Matches' tab (LOCKED Masterlist ID(s) -> SellUp SKU ID) plus any
    'Linked' rows in 'New Masterlist SKUs'. Returns {stock_id: sellup_sku_id}.
    This lets accumulated decisions persist even after they move into Locked."""
    import openpyxl
    wb = openpyxl.load_workbook(path_or_stream, data_only=True)
    out = {}

    if "Locked Matches" in wb.sheetnames:
        ws = wb["Locked Matches"]
        hdr = [str(c.value) for c in ws[1]]
        try:
            ci = hdr.index("LOCKED Masterlist ID(s)")
            cs = hdr.index("SellUp SKU ID")
        except ValueError:
            ci = cs = None
        if ci is not None and cs is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                sku = str(r[cs]).strip() if r[cs] not in (None, "") else ""
                if not sku or r[ci] in (None, ""):
                    continue
                for part in str(r[ci]).split(","):
                    mid = _norm_id(part)
                    if mid:
                        out.setdefault(mid, sku)

    if "New Masterlist SKUs" in wb.sheetnames:
        ws = wb["New Masterlist SKUs"]
        hdr = [str(c.value) for c in ws[1]]

        def idx(name):
            return hdr.index(name) if name in hdr else None
        ci, cd, cl = (idx("Masterlist Stock Type ID"),
                      idx("Reviewer Decision"), idx("Link to SellUp SKU ID"))
        if ci is not None and cl is not None:
            for r in ws.iter_rows(min_row=2, values_only=True):
                mid = _norm_id(r[ci])
                sku = str(r[cl]).strip() if cl is not None and r[cl] not in (None, "") else ""
                dec = str(r[cd]).strip().lower() if cd is not None and r[cd] else ""
                if mid and sku and ("link" in dec or not dec):
                    out[mid] = sku          # newest link wins
    return out


def read_prior_decisions(path_or_stream) -> dict:
    """Read reviewer decisions from a prior registry's New Masterlist SKUs tab.
    Returns {stock_id: {"decision": str, "link": str}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path_or_stream, data_only=True)
    if "New Masterlist SKUs" not in wb.sheetnames:
        return {}
    ws = wb["New Masterlist SKUs"]
    hdr = [str(c.value) for c in ws[1]]

    def idx(name):
        return hdr.index(name) if name in hdr else None
    ci, cd, cl = (idx("Masterlist Stock Type ID"),
                  idx("Reviewer Decision"), idx("Link to SellUp SKU ID"))
    out = {}
    if ci is None:
        return out
    for r in ws.iter_rows(min_row=2, values_only=True):
        sid = _norm_id(r[ci])
        if not sid:
            continue
        dec = str(r[cd]).strip() if cd is not None and r[cd] not in (None, "") else ""
        link = str(r[cl]).strip() if cl is not None and r[cl] not in (None, "") else ""
        if dec or link:
            out[sid] = {"decision": dec, "link": link}
    return out


_TAB_SIZE_RE = re.compile(r"\b(9\.7|10\.2|10\.5|10\.9|12\.9|13\.0|11\.0|13|11)\b")


def tablet_size(text: str) -> str:
    """Tablet screen size (inches) e.g. '12.9', '11'. '' if none."""
    mm = _TAB_SIZE_RE.search(_up(text))
    if not mm:
        return ""
    return {"11.0": "11", "13.0": "13"}.get(mm.group(1), mm.group(1))


def suggest_sellup(m: "MasterRow", sellup_by_brand: dict) -> tuple:
    """Best-guess SellUp row for an unmatched Masterlist item, scored on all
    five dimensions: Model + Storage + Connectivity + Colour + Condition
    (Condition = the SellUp listing prices the Masterlist row's grade).

    Same-generation/tier products get a strong bonus (so a Pro 1 won't be
    suggested for a Pro 2), but the signature is not a hard gate — this lets
    divergently-named items (e.g. iPad 'PRO 12.9 4 GEN 2020' vs
    'iPad Pro 12.9 (4th Gen)') still surface a reviewable suggestion.
    Returns (SellUpRow|None, score)."""
    m_sig = model_signature(m.model_base)
    m_tsz = tablet_size(m.raw_model) if m.dtype == "tablet" else ""
    best, best_score = None, -1.0
    for s in sellup_by_brand.get(m.brand, []):
        if s.dtype != m.dtype:
            continue  # never suggest an iPhone for an iPad, a watch for audio…
        same_sig = model_signature(s.model_base) == m_sig
        # Phones / watches / audio: strict generation+tier gate (no cross-gen).
        # Tablets: divergent naming, so signature is a strong bonus, not a gate.
        if m.dtype != "tablet" and not same_sig:
            continue
        sc = fuzz.token_set_ratio(m.model_base, s.model_base)
        if same_sig:
            sc += 25                                   # same gen/tier — strong
        # Storage — a wrong storage is a different product.
        if m.storage or s.storage:
            sc += 10 if (m.storage and m.storage == s.storage) else -40
        # Tablet screen size — 11" vs 12.9" are different products.
        if m.dtype == "tablet" and s.dtype == "tablet":
            s_tsz = tablet_size(s.raw_model)
            if m_tsz and s_tsz:
                sc += 10 if m_tsz == s_tsz else -50
        # Colour
        if m.colour and s.colour:
            sc += 10 if m.colour == s.colour else -15
        # Connectivity (tablets / watches)
        if m.connectivity and s.connectivity:
            sc += 8 if m.connectivity == s.connectivity else -20
        # Watch case size
        if m.size_mm and s.size_mm:
            sc += 8 if m.size_mm == s.size_mm else -20
        # Condition — does this SellUp listing sell the Masterlist row's grade?
        sc += 12 if s.price_active(m.condition) else -20
        if sc > best_score:
            best, best_score = s, sc
    return best, best_score


def reconcile_decisions(results, unmatched_master, rows_by_sheet,
                        manual_decisions=None):
    """Apply freebies rule + carried-forward reviewer decisions.

    Returns (results_augmented, buckets) where buckets maps
    'new' / 'not_yet' / 'not_selling' -> list[(MasterRow, reason)].
    Manual 'Linked' rows become synced matches appended to results.
    """
    manual = manual_decisions or {}
    sellup_by_sku = {s.sku_id: s for rows in rows_by_sheet.values() for s in rows}
    buckets = {"new": [], "not_yet": [], "not_selling": []}
    manual_links = []
    for m in unmatched_master:
        if is_freebie(m):
            buckets["not_selling"].append((m, "Freebie (auto)"))
            continue
        if is_telco(m):
            buckets["not_selling"].append((m, "Telco — Primary only (auto)"))
            continue
        info = manual.get(m.stock_id, {})
        dec = (info.get("decision") or "").strip()
        link = (info.get("link") or "").strip()
        d = dec.lower()
        if "link" in d and link:
            s = sellup_by_sku.get(link)
            if s:
                manual_links.append((m, s))
            else:
                buckets["new"].append((m, f"Link SKU '{link}' not found"))
        elif "not selling" in d:
            buckets["not_selling"].append((m, dec))
        elif "not on sellup" in d:
            buckets["not_yet"].append((m, dec))
        elif "skip" in d:
            buckets["not_selling"].append((m, dec))
        else:
            buckets["new"].append((m, ""))

    agg = defaultdict(lambda: [None, 0, []])
    for m, s in manual_links:
        cell = agg[(s.sku_id, m.condition)]
        cell[0] = s
        cell[2].append(m)
        if s.price_active(m.condition):
            cell[1] += m.qty
    extra = [MatchResult(s, cond, mss, qty, "High", 100.0, note="manual link")
             for (sku, cond), (s, qty, mss) in agg.items()]
    return results + extra, buckets


def build_review_workbook(results: list[MatchResult],
                          unmatched_master: list[MasterRow],
                          master: list[MasterRow],
                          rows_by_sheet: dict[str, list[SellUpRow]],
                          applied_note: str = "",
                          manual_decisions: dict | None = None):
    """Build the SellUp Match Review registry workbook, styled like Shopee's.

    manual_decisions: {stock_id: {"decision","link"}} carried forward from a
    prior reviewed registry. Freebies auto-route to "Not Selling in SellUp".
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    results, buckets = reconcile_decisions(
        results, unmatched_master, rows_by_sheet, manual_decisions)
    b_new = buckets["new"]
    b_not_yet = buckets["not_yet"]
    b_not_selling = buckets["not_selling"]

    wb = Workbook()
    wb.remove(wb.active)

    white = Font(bold=True, color="FFFFFF")
    black = Font(bold=True, color="000000")
    body = Font(size=10)
    navy_fill = PatternFill("solid", fgColor=CLR_NAVY)
    peach_fill = PatternFill("solid", fgColor=CLR_PEACH)
    gold_fill = PatternFill("solid", fgColor=CLR_GOLD)
    center = Alignment(vertical="center", wrap_text=True)

    def make_sheet(name, headers, colours, widths):
        """headers: list[str]; colours: list of 'navy'|'peach'|'gold' per column."""
        ws = wb.create_sheet(name)
        ws.append(headers)
        fills = {"navy": navy_fill, "peach": peach_fill, "gold": gold_fill}
        fonts = {"navy": white, "peach": black, "gold": black}
        for idx, (cell, kind) in enumerate(zip(ws[1], colours)):
            cell.fill = fills[kind]
            cell.font = fonts[kind]
            cell.alignment = center
        for i, w in enumerate(widths):
            ws.column_dimensions[ws.cell(1, i + 1).column_letter].width = w
        ws.freeze_panes = "A2"
        return ws

    def finalize(ws):
        ws.auto_filter.ref = f"A1:{ws.cell(1, ws.max_column).column_letter}1"
        for row in ws.iter_rows(min_row=2):
            for c in row:
                if c.font is None or not c.font.bold:
                    c.font = body

    def add_dropdown(ws, col_idx, options):
        """Attach a data-validation dropdown to a column (rows 2..end)."""
        if ws.max_row < 2:
            return
        letter = ws.cell(1, col_idx).column_letter
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(options) + '"',
                            allow_blank=True)
        dv.promptTitle = "Reviewer Decision"
        dv.prompt = "Pick one: " + " / ".join(options)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{ws.max_row}")

    def _distinct(vals):
        return " | ".join(sorted({str(v) for v in vals if str(v).strip()}))

    def ml_model_color(rows):
        # "id:MODEL|COLOR" like Shopee's "30362:Z FOLD 7 ...|SILVER SHADOW"
        return " ; ".join(f"{m.stock_id}:{m.raw_model}|{m.raw_colour or m.colour}"
                          for m in rows)

    # ========================= Locked Matches =========================
    lock_hdr = ["#", "SellUp Sheet", "SellUp SKU ID", "SellUp Model", "Storage",
                "Connectivity", "SellUp Colour", "Condition Column",
                "LOCKED Masterlist ID(s)", "ML Category", "ML Model(s)|Color",
                "ML Available Qty", "Target Qty", "# SKUs"]
    lock_clr = (["navy"] + ["peach"] * 7 + ["gold"] * 5 + ["navy"])
    lock_w = [5, 14, 16, 22, 9, 12, 15, 18, 22, 12, 55, 14, 11, 7]
    ws_lock = make_sheet("Locked Matches", lock_hdr, lock_clr, lock_w)
    n = 0
    for r in results:
        if r.confidence in ("Exact", "High"):
            n += 1
            ws_lock.append([
                n, r.sellup.sheet, r.sellup.sku_id,
                (r.sellup.raw_model or r.sellup.model_base),
                r.sellup.storage, r.sellup.connectivity,
                (r.sellup.raw_colour or r.sellup.colour),
                r.condition, ", ".join(m.stock_id for m in r.master_rows),
                _distinct(m.category for m in r.master_rows),
                ml_model_color(r.master_rows),
                sum(m.qty for m in r.master_rows), r.qty, len(r.master_rows),
            ])
    finalize(ws_lock)

    # ========================= New Masterlist SKUs =========================
    # Masterlist items (MM stock) that are NOT matched to any SellUp listing
    # yet -> candidates to list / link on SellUp.
    nm_hdr = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model",
              "Color", "Available Qty", "Routed Condition",
              "Suggested SellUp SKU ID", "Suggested Model", "Suggested Storage",
              "Suggested Connectivity", "Suggested Colour",
              "Suggested Condition(s)", "Match %",
              "Link to SellUp SKU ID", "Reviewer Decision", "Notes"]
    nm_clr = (["navy"] + ["peach"] * 7 + ["gold"] * 10)
    nm_w = [5, 20, 10, 12, 34, 16, 12, 18,
            20, 26, 10, 14, 16, 26, 8, 20, 20, 22]
    ws_nm = make_sheet("New Masterlist SKUs", nm_hdr, nm_clr, nm_w)
    sellup_by_brand = defaultdict(list)
    for rows in rows_by_sheet.values():
        for s in rows:
            sellup_by_brand[s.brand].append(s)

    def active_conditions(s):
        return ", ".join(c for c in CONDITIONS if s.price_active(c))

    n = 0
    for m, note in b_new:
        n += 1
        s, score = suggest_sellup(m, sellup_by_brand)
        if s and score >= 45:
            sug = [s.sku_id, (s.raw_model or s.model_base), s.storage,
                   s.connectivity, (s.raw_colour or s.colour),
                   active_conditions(s), round(min(score, 100.0), 0)]
        else:
            sug = [None] * 7
        ws_nm.append([n, m.stock_id, m.category, m.brand, m.raw_model,
                      m.raw_colour or m.colour, m.qty, m.condition,
                      *sug, None, None, note or None])
    finalize(ws_nm)
    # Reviewer Decision (col 17) dropdown.
    add_dropdown(ws_nm, 17, NM_DECISIONS)

    # ========================= Match Review =========================
    # Every active-priced SellUp SKU that did NOT lock to the Masterlist:
    #   - "None":   no Masterlist match at all (live on SellUp, no MM stock)
    #   - "Review": a weak fuzzy suggestion exists (confirm or skip)
    mr_hdr = ["#", "SellUp Sheet", "SellUp SKU ID", "SellUp Model", "Storage",
              "Connectivity", "SellUp Colour", "Condition Column", "Current Qty",
              "Match Status", "Suggested Masterlist ID", "ML Category",
              "ML Model|Color", "Score", "Reviewer Decision", "Notes"]
    mr_clr = (["navy"] + ["peach"] * 8 + ["gold"] * 7)
    mr_w = [5, 14, 16, 22, 9, 12, 15, 18, 11, 16, 20, 12, 45, 7, 22, 24]
    ws_mr = make_sheet("Match Review", mr_hdr, mr_clr, mr_w)
    n = 0
    for r in results:
        if r.confidence not in ("None", "Review"):
            continue
        n += 1
        cur = r.sellup.qtys.get(r.condition)
        if r.confidence == "Review":
            status = "Fuzzy suggestion"
            sug_id = ", ".join(m.stock_id for m in r.master_rows)
            sug_cat = _distinct(m.category for m in r.master_rows)
            sug_mc = ml_model_color(r.master_rows)
            score = round(r.score, 1)
            decision = "Confirm? (LOCKED / SKIP)"
        else:  # None
            status = "No Masterlist match"
            sug_id = sug_cat = sug_mc = ""
            score = ""
            decision = "SKIP / DELIST?"
        ws_mr.append([
            n, r.sellup.sheet, r.sellup.sku_id,
            (r.sellup.raw_model or r.sellup.model_base),
            r.sellup.storage, r.sellup.connectivity,
            (r.sellup.raw_colour or r.sellup.colour), r.condition, cur,
            status, sug_id, sug_cat, sug_mc, score, decision, None,
        ])
    finalize(ws_mr)

    # ========================= Skipped (No Price) =========================
    sk_hdr = ["#", "SellUp Sheet", "SellUp SKU ID", "SellUp Model", "Storage",
              "SellUp Colour", "Reason"]
    sk_clr = (["navy"] + ["peach"] * 5 + ["navy"])
    sk_w = [5, 14, 16, 24, 9, 15, 24]
    ws_sk = make_sheet("Skipped (No Price)", sk_hdr, sk_clr, sk_w)
    n = 0
    for _sheet, rows in rows_by_sheet.items():
        for s in rows:
            if not any(s.price_active(c) for c in CONDITIONS):
                n += 1
                ws_sk.append([n, s.sheet, s.sku_id,
                              (s.raw_model or s.model_base), s.storage,
                              (s.raw_colour or s.colour), "No active price"])
    finalize(ws_sk)

    # ========================= Not Selling in SellUp =========================
    # Freebies (auto) + items the reviewer marked as not-selling / skipped.
    ns_hdr = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model",
              "Color", "Available Qty", "Reason"]
    ns_clr = (["navy"] + ["peach"] * 6 + ["gold"])
    ns_w = [5, 20, 10, 12, 40, 16, 12, 22]
    ws_ns = make_sheet("Not Selling in SellUp", ns_hdr, ns_clr, ns_w)
    n = 0
    for m, reason in b_not_selling:
        n += 1
        ws_ns.append([n, m.stock_id, m.category, m.brand, m.raw_model,
                      m.raw_colour or m.colour, m.qty, reason])
    finalize(ws_ns)

    # ========================= Not on SellUp Yet =========================
    # Masterlist items the reviewer marked "Not on SellUp Yet".
    ny_hdr = ["#", "Masterlist Stock Type ID", "Category", "Brand", "Model",
              "Color", "Available Qty", "Routed Condition"]
    ny_clr = (["navy"] + ["peach"] * 6 + ["gold"])
    ny_w = [5, 20, 10, 12, 40, 16, 12, 18]
    ws_ny = make_sheet("Not on SellUp Yet", ny_hdr, ny_clr, ny_w)
    n = 0
    for m, _reason in b_not_yet:
        n += 1
        ws_ny.append([n, m.stock_id, m.category, m.brand, m.raw_model,
                      m.raw_colour or m.colour, m.qty, m.condition])
    finalize(ws_ny)

    # ========================= Summary (first tab) =========================
    stats = summarise(results)
    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum["B2"] = "SellUp Stock Bulk Update — Match Registry"
    ws_sum["B2"].font = Font(bold=True, size=15, color=CLR_NAVY)
    ws_sum["B4"] = "Category"
    ws_sum["C4"] = "Count"
    for coord in ("B4", "C4"):
        ws_sum[coord].font = white
        ws_sum[coord].fill = navy_fill
    summary_rows = [
        ("Locked Matches (sync stock)", stats.get("Exact", 0) + stats.get("High", 0)),
        ("New Masterlist SKUs (to review)", ws_nm.max_row - 1),
        ("Match Review (SellUp SKUs, no MM match)",
         stats.get("None", 0) + stats.get("Review", 0)),
        ("Skipped (No active price)", ws_sk.max_row - 1),
        ("Not Selling in SellUp (incl. freebies)", ws_ns.max_row - 1),
        ("Not on SellUp Yet", ws_ny.max_row - 1),
    ]
    row = 5
    for label, count in summary_rows:
        ws_sum[f"B{row}"] = label
        ws_sum[f"C{row}"] = count
        ws_sum[f"B{row}"].font = body
        ws_sum[f"C{row}"].font = body
        row += 1
    row += 1
    ws_sum[f"B{row}"] = ("Routing:  New+NA → New (Not Activated)  |  "
                         "New+A → New (Activated)  |  Used → Excellent")
    ws_sum[f"B{row}"].font = Font(size=10, italic=True, color=CLR_GREY)
    row += 1
    ws_sum[f"B{row}"] = ("Write rule:  Qty only, and only where that "
                         "condition's Price cell > 0.")
    ws_sum[f"B{row}"].font = Font(size=10, italic=True, color=CLR_GREY)
    if applied_note:
        row += 1
        ws_sum[f"B{row}"] = applied_note
        ws_sum[f"B{row}"].font = Font(size=10, color=CLR_GREY)
    ws_sum.column_dimensions["A"].width = 3
    ws_sum.column_dimensions["B"].width = 42
    ws_sum.column_dimensions["C"].width = 10

    return wb
